# author:ling
# date:2021.11.11

import re
import pandas as pd
import urllib.request
import chardet
import random
from openpyxl import load_workbook
import time
from selenium import webdriver
from scrapy import Selector

class Nowcoder(object):

    def __init__(self):

        # 读入xlsx文件中的链接
        data = pd.read_excel('nowcoder.xlsx')
        self.urls = data.iloc[:,1]

        self.salary = []  # 薪资
        self.area = []  # 地区
        self.education = []  # 学历

        self.com_name = []  #公司名称
        self.com_tag=[]  # 公司类别


        self.job = []  # 岗位职责
        self.need = []  # 岗位需求

        self.timeout = 2
        self.filename = 'nowcoder.xlsx'

        self.getInfo()
        self.saveInfo()

    def subStr(self,st):
        '''去除单引号'''
        subStr = []
        for s in st:
            s = s.strip("'")
            subStr.append(s)
        return subStr

    def getInfo(self):
        for i in range(len(self.urls)):
            # 获取页面html

            # driver设置
            chrome_options = webdriver.ChromeOptions()
            chrome_options.add_argument('--headless')
            driver = webdriver.Chrome(options=chrome_options)
            # 浏览器页面
            driver.get(self.urls[i])

            # 异步加载的数据需要等待时长获取信息
            waittime = 1
            loadmore = False
            time.sleep(waittime)
            if loadmore:
                while True:
                    try:
                        next_button = driver.find_element_by_class_name("more")
                        next_button.click()
                        time.sleep(waittime)
                    except:
                        break
            # 获取信息
            html = driver.page_source

            salary = re.findall(r'<div data-v-513b791a="" class="salary mlr-4">(.*?)</div>', html)
            salary = self.subStr(salary)

            area = re.findall(r'tabindex="0">(.*?)</span>', html)
            area = self.subStr(area)

            education = re.findall(r'</div> <span data-v-513b791a="">(.*?)</span>', html)
            education = self.subStr(education)

            com_name = re.findall(r'class="mt-3 mb-1 hover-green fs-20">(.*?)</div>', html)
            com_name = self.subStr(com_name)

            com_tag = re.findall(r'class="company-tags"><span data-v-243f35ea="">(.*?)</span>',html)
            com_tag = self.subStr(com_tag)

            # job和need的re索引有问题，因此该部分换成scrapy爬虫
            selector = Selector(text=html)
            job = selector.xpath('//div[@class="ptb-2 pre-line"]/text()').extract()[0]
            job = job.replace("\n","")
            need = selector.xpath('//div[@class="ptb-2 pre-line"]/text()').extract()[-1]
            need = need.replace("\n", "")

            time.sleep(self.timeout)
            print("%d finish" % (i+1))

            # 填入数组之中
            self.salary.append(salary)
            self.area.append(area)
            self.education.append(education)

            self.com_name.append(com_name)
            self.com_tag.append(com_tag)

            self.job.append(job)
            self.need.append(need)

            driver.quit()

    def saveInfo(self):
        wb = load_workbook(self.filename)
        wb.create_sheet('info',1)
        ws = wb['info']

        ws.append(['薪水', '地区', '学历', '公司名称', '公司类别', '岗位职责', '岗位需求'])

        for i in range(len(self.salary)):
            line = [str(self.salary[i]), str(self.area[i]), str(self.education[i]), str(self.com_name[i]), str(self.com_tag[i]),
                    str(self.job[i]), str(self.need[i])]
            ws.append(line)
            # time.sleep(self.timeout)

        wb.save(self.filename)
        print("save!")


if __name__ == '__main__':
    nowcoder = Nowcoder()
