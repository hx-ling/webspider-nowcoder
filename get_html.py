# author:ling
# date: 2021.11.11

from selenium import webdriver
import time
import re
from openpyxl import load_workbook
from openpyxl import Workbook

'''
步骤：
1、获取一页内容
2、直接re库找到链接内容
3、写入xlsx文件中
4、点击下一页
5、循环15次

组件：
1、2、3、4写为一个函数
5作为main函数调用
'''

def One_page(driver):

    '''获取一页内容'''

    # 异步加载的数据需要等待时长获取信息
    waittime = 2
    loadmore = False
    time.sleep(waittime)
    if loadmore:
        while True:
            try:
                next_button = driver.find_element_by_class_name("more")
                next_button.click()
                time.sleep(waittime)
            except:
                break
    # 获取信息
    html = driver.page_source

    # '''写入html文件，方便查看'''
    # get_html = "nowcoder.html"
    # # 打开文件，准备写入
    # f = open(get_html, 'wb')
    # time.sleep(2)  # 保证浏览器响应成功后再进行下一步操作
    # # 写入文件
    # f.write(html.encode("utf-8", "ignore"))  # 忽略非法字符
    # print('写入成功')
    # # 关闭文件
    # f.close()

    # 查看
    # print(html)
    return html

def getInfo(html):

    '''re获取信息'''

    # 获取信息
    job_names = re.findall(r'<span data-v-4f4cbcac="" class="job-name">(.*?)</span>', html)  # 职业名称
    links = re.findall(r'<a data-v-4f4cbcac="" href="https(.*?)" target="_blank" class="job-message-boxs">', html)  # 职业信息

    time.sleep(3)
    print("finish")

    return job_names,links

def saveData(jobNames,links):

    '''存储数据'''

    filename = './nowcoder.xlsx'

    wb = Workbook()
    wb.create_sheet('链接',0)
    ws = wb['链接']

    # workbook = xlrd.open_workbook(self.filename)
    # worksheet = workbook.sheet_by_index(0)
    # cell_value = worksheet.cell(1,0)

    ws.append(['职位名称', '链接'])

    for i in range(len(links)):
        line = [jobNames[i],links[i]]
        ws.append(line)
        # time.sleep(self.timeout)

    wb.save(filename)
    print("save")


if __name__ == '__main__':
    # driver设置
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument('--headless')
    driver = webdriver.Chrome(options=chrome_options)
    # driver = webdriver.Chrome()
    # 浏览器页面
    driver.get('https://www.nowcoder.com/school/jobs?scheduleTab=1&pageSource=5001&firstScroll=true&careerJob=11006')

    jobNames = []
    Links = []

    for i in range(25):
        # 获取单页面
        one_html = One_page(driver)
        # re索取
        job_names,links = getInfo(one_html)
        for n in range(len(links)):
            jobNames.append(job_names[n])
            Links.append('https'+links[n])
        # 点击下一页
        driver.find_element_by_xpath("//button[@class='btn-next']").click()

    # 存储信息
    saveData(jobNames, Links)

    driver.quit()

