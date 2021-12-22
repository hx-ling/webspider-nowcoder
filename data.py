# author:ling
# date:2021.11.12

from openpyxl import load_workbook
import numpy as np
import pandas as pd


'''
对数据做初步清洗，去除符号和无用信息
'''
# open workbook
wb = load_workbook('nowcoder.xlsx')
ws = wb['info']

rows = ws.max_row
cols = ws.max_column

def sub_str(s):
    s = s.replace('[','')
    s = s.replace(']','')
    s = s.strip("'")
    return s

for row in range(1,rows+1):
    ws.cell(row,1).value = sub_str(ws.cell(row,1).value)
    ws.cell(row,3).value = sub_str(ws.cell(row,3).value)
    ws.cell(row,4).value = sub_str(ws.cell(row,4).value)
    ws.cell(row,5).value = sub_str(ws.cell(row,5).value)

    area = ws.cell(row,2).value
    area = area.split(",")
    area = area[0]
    area = area.replace('[', '')
    area = area.replace(']', '')
    area = area.replace("'","")

    ws.cell(row,2).value = area

# 保存
wb.save('./nowcoder.xlsx')
print("save!")