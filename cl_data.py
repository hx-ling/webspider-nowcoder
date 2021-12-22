#author:ling
#date:2021.11.15

'''数据清洗'''

from openpyxl import load_workbook
import numpy as np

wb = load_workbook('nowcoder.xlsx')
ws = wb['info']

rows = ws.max_row
cols = ws.max_column

ws.cell(1,8).value = "最低薪资"
ws.cell(1,9).value = "最高薪资"


for row in range(2,rows+1):
    s = ws.cell(row,1).value

    if s == '薪资面议':
        min_s = 0
        max_s = 0
    else:
        min_s = int(s.split("-")[0])
        max_s = s.split("-")[1]
        max_s = int(max_s.split("K")[0])

    ws.cell(row,8).value = min_s
    ws.cell(row,9).value = max_s

wb.save('nowcoder.xlsx')
print("finish!")

